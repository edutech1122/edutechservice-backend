"""
Excel workbook generation -- generalized from the Diploma-in-Pharmacy
prototype (`build_pharmacy_sample.py`) to work for any course/scheme/
pattern. Sheet layouts, styling, and the analytics rules below were
verified against the user's own real sample screenshots (see
MSBTE_Result_Analysis_Architecture.md addenda) -- this module intentionally
writes plain computed numbers rather than Excel formulas, per the user's
explicit instruction for this workbook's sheets.

Produces, per selected course:
  - one Internal / External / Total sheet per stage (year or semester)
  - one combined "<Scheme><stage-count>" style J11 "Result Analysis" sheet
  - one J12 "Sessional TM vs Annual TH analysis" sheet
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .marks import analytics, HEAD_MAX

A4_PORTRAIT_UNITS = 100
A4_LANDSCAPE_UNITS = 148


def apply_print_setup(ws, last_col, last_row_for_header=1):
    total_width = sum(
        (ws.column_dimensions[get_column_letter(c)].width or 8.43)
        for c in range(1, last_col + 1)
    )
    portrait = total_width <= A4_PORTRAIT_UNITS
    ws.page_setup.orientation = 'portrait' if portrait else 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    ws.print_title_rows = f'{last_row_for_header}:{last_row_for_header}'
    ws.print_area = f'A1:{get_column_letter(last_col)}{ws.max_row}'
    return portrait


HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
BAND_FILL = PatternFill('solid', fgColor='DCE6F1')
FOOT_FILL = PatternFill('solid', fgColor='2E5395')
FOOT_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Calibri', size=10)
BOLD = Font(name='Calibri', bold=True, size=10)
THIN = Side(style='thin', color='B7C6D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
RANK_FILL = {1: PatternFill('solid', fgColor='2E7D32'),
             2: PatternFill('solid', fgColor='66BB6A'),
             3: PatternFill('solid', fgColor='C8E6C9')}
BANNER_FILL = PatternFill('solid', fgColor='1F3864')
SUBTITLE_FILL = PatternFill('solid', fgColor='DCE6F1')
DATA_FILL = PatternFill('solid', fgColor='FFF7CC')
BAND_FILL_2 = PatternFill('solid', fgColor='F2F2F2')
SECTION_FILL = PatternFill('solid', fgColor='2E5395')
SIGN_FONT = Font(name='Calibri', italic=True, bold=True, color='1F3864', size=11)
DIFF_PINK = PatternFill('solid', fgColor='F8CBAD')
DIFF_GREEN = PatternFill('solid', fgColor='C6E0B4')


def compute_academic_year(session_text: str) -> str:
    """SUMMER 2026 -> 2025-2026; WINTER 2026 -> 2026-2027 (June-May Indian
    academic calendar assumption -- documented in the Notes sheet)."""
    m = re.search(r'(SUMMER|WINTER)\s+(\d{4})', session_text.upper())
    if not m:
        return session_text
    kind, year = m.group(1), int(m.group(2))
    return f'{year - 1}-{year}' if kind == 'SUMMER' else f'{year}-{year + 1}'


def write_sheet(wb, title, block, component, with_rank_cols=False, code_map=None):
    code_map = code_map or {}
    ws = wb.create_sheet(title[:31])
    cols = block['cols']
    students = block['students']

    base_headers = ['Sr. No.', 'Seat No.', 'Enrollment No.', 'Name']
    extra_headers = ['Total Marks', 'Percentage', 'Rank', 'Result'] if with_rank_cols else []
    n_base = len(base_headers)
    n_subj = len(cols)

    for i, h in enumerate(base_headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    for j, c in enumerate(cols, start=1):
        col = n_base + j
        cell = ws.cell(row=1, column=col, value=c['subject'])
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
    for k, h in enumerate(extra_headers, start=1):
        col = n_base + n_subj + k
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER

    ranks = {}
    if with_rank_cols:
        totals = []
        for s in students:
            tot = sum(v['tot'] for v in s['heads'].values() if v['applicable'] and v['tot'] is not None)
            totals.append((s['seat'], tot))
        sorted_totals = sorted(totals, key=lambda x: -x[1])
        rank = 0; prev = None; seen = 0
        for seat, tot in sorted_totals:
            seen += 1
            if tot != prev:
                rank = seen
                prev = tot
            ranks[seat] = (rank, tot)

    r = 2
    for i, s in enumerate(students, start=1):
        row_fill = BAND_FILL if i % 2 == 0 else None
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=s['seat'])
        ws.cell(row=r, column=3, value=s['enroll'])
        ws.cell(row=r, column=4, value=s['name'])
        for j, c in enumerate(cols, start=1):
            v = s['heads'][c['n']][component]
            disp = v if v is not None else ('OPT' if not s['heads'][c['n']]['applicable'] else '')
            ws.cell(row=r, column=n_base + j, value=disp)
        if with_rank_cols:
            rank, tot = ranks[s['seat']]
            pct = round(tot / (n_subj * HEAD_MAX) * 100, 2) if n_subj else 0
            ws.cell(row=r, column=n_base + n_subj + 1, value=tot)
            ws.cell(row=r, column=n_base + n_subj + 2, value=pct)
            ws.cell(row=r, column=n_base + n_subj + 3, value=rank)
            ws.cell(row=r, column=n_base + n_subj + 4, value=s.get('result') or '')
        for col in range(1, n_base + n_subj + len(extra_headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if col != 4 else Alignment(horizontal='left', vertical='center')
            if row_fill:
                cell.fill = row_fill
        if with_rank_cols:
            rank = ranks[s['seat']][0]
            if rank in RANK_FILL:
                for col in range(1, n_base + n_subj + len(extra_headers) + 1):
                    ws.cell(row=r, column=col).fill = RANK_FILL[rank]
        r += 1

    stats = analytics(students, cols, component)

    def foot_row(label, values, fmt=None):
        nonlocal r
        c0 = ws.cell(row=r, column=4, value=label); c0.font = FOOT_FONT; c0.fill = FOOT_FILL; c0.border = BORDER
        for col in range(1, 4):
            cc = ws.cell(row=r, column=col); cc.fill = FOOT_FILL; cc.border = BORDER
        for j, c in enumerate(cols, start=1):
            val = values.get(c['n'])
            cell = ws.cell(row=r, column=n_base + j, value=val)
            cell.font = FOOT_FONT; cell.fill = FOOT_FILL; cell.alignment = CENTER; cell.border = BORDER
            if fmt: cell.number_format = fmt
        for col in range(n_base + n_subj + 1, n_base + n_subj + len(extra_headers) + 1):
            cc = ws.cell(row=r, column=col); cc.fill = FOOT_FILL; cc.border = BORDER
        r += 1

    foot_row('Subjects', {c['n']: c['subject'] for c in cols})
    foot_row('Lowest Marks Obtained', {c['n']: stats[c['n']]['low'] for c in cols})
    foot_row('Highest Marks Obtained', {c['n']: stats[c['n']]['high'] for c in cols})
    foot_row('No. of Students appeared', {c['n']: stats[c['n']]['appeared'] for c in cols})
    foot_row('No. of Students passed', {c['n']: stats[c['n']]['passed'] for c in cols})
    foot_row('Percentage passed', {c['n']: stats[c['n']]['pct_passed'] / 100 for c in cols}, fmt='0%')
    foot_row('Percentage of Students above 60%', {c['n']: stats[c['n']]['pct_above60'] / 100 for c in cols}, fmt='0%')

    for col in range(1, n_base + n_subj + len(extra_headers) + 1):
        cc = ws.cell(row=r, column=col); cc.fill = FOOT_FILL; cc.border = BORDER
    if n_subj:
        first_subj_col = n_base + 1
        last_subj_col = n_base + n_subj
        ws.merge_cells(start_row=r, start_column=first_subj_col, end_row=r, end_column=last_subj_col)
        hdr = ws.cell(row=r, column=first_subj_col, value='Score index')
        hdr.font = FOOT_FONT; hdr.alignment = CENTER
    r += 1
    foot_row('Total marks secured by all student', {c['n']: stats[c['n']]['total_secured'] for c in cols})
    foot_row('Score index', {c['n']: stats[c['n']]['score_index'] for c in cols})

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 27
    for j in range(1, n_subj + 1):
        ws.column_dimensions[get_column_letter(n_base + j)].width = 6.5
    extra_widths = {'Total Marks': 8.5, 'Percentage': 8.5, 'Rank': 6, 'Result': 15}
    for k, h in enumerate(extra_headers, start=1):
        ws.column_dimensions[get_column_letter(n_base + n_subj + k)].width = extra_widths.get(h, 9)
    ws.freeze_panes = 'E2'
    ws.sheet_view.showGridLines = False
    apply_print_setup(ws, n_base + n_subj + len(extra_headers))
    return ws


def build_result_analysis_sheet(wb, sheet_title, sections, pattern_label, institute_name,
                                 session_text, academic_year, coordinator_name,
                                 coordinator_role, principal_name, code_map=None):
    """J11-style consolidated proforma: one Annual (=Total-stats) row and one
    Sessional (=Internal-stats) row per subject, grouped under a divider row
    per stage/section. `sections` is an ordered list of (section_label,
    block)."""
    code_map = code_map or {}
    form_code = f"{sections[0][1]['scheme']}11"
    ws = wb.create_sheet(f'{sheet_title} {form_code}'[:31])

    headers = ['Sr.\nNo.', 'Course Name', 'Course Code', 'Passing Heads',
               'Marks obtained\nLowest', 'Marks obtained\nHighest',
               'No. of Students\nappeared', 'No. of students\nPassed',
               '% Pass', '% of students\nabove 60%']
    n = len(headers)

    c = ws.cell(row=1, column=n, value=form_code)
    c.font = Font(name='Calibri', italic=True, bold=True, size=11)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
    c = ws.cell(row=3, column=1, value='Maharashtra State Board of Technical Education, Mumbai')
    c.font = Font(name='Calibri', bold=True, size=16, color='FFFFFF'); c.fill = BANNER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 26

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n)
    c = ws.cell(row=4, column=1, value=f'RESULT ANALYSIS OF {pattern_label} EXAMINATION')
    c.font = Font(name='Calibri', bold=True, size=13, color='1F3864'); c.fill = SUBTITLE_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n)
    c = ws.cell(row=5, column=1, value=institute_name)
    c.font = Font(name='Calibri', bold=True, size=12); c.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    ws.cell(row=6, column=1, value='Academic Year:').font = BOLD
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=5)
    ws.cell(row=6, column=3, value=academic_year)
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7)
    ws.cell(row=6, column=6, value='Examination:').font = BOLD
    ws.merge_cells(start_row=6, start_column=8, end_row=6, end_column=n)
    ws.cell(row=6, column=8, value=session_text)
    ws.row_dimensions[6].height = 18
    for col in range(1, n + 1):
        ws.cell(row=6, column=col).border = BORDER
        ws.cell(row=6, column=col).alignment = Alignment(horizontal='left', vertical='center')

    hr = 8
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[hr].height = 48

    r = hr + 1
    for section_label, block in sections:
        cols = block['cols']
        students = block['students']
        tot_stats = analytics(students, cols, 'tot')
        inn_stats = analytics(students, cols, 'inn')

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        c = ws.cell(row=r, column=1, value=section_label.upper())
        c.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        c.fill = SECTION_FILL; c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1

        for i, c in enumerate(cols, start=1):
            band = BAND_FILL_2 if i % 2 == 0 else None
            for comp_label, stats in (('Annual', tot_stats), ('Sessional', inn_stats)):
                s = stats[c['n']]
                row_vals = [i if comp_label == 'Annual' else None,
                            c['subject'] if comp_label == 'Annual' else None,
                            code_map.get(c['subject'], '') if comp_label == 'Annual' else None,
                            f"{c['head']}-{comp_label}",
                            s['low'], s['high'], s['appeared'], s['passed'],
                            s['pct_passed'], s['pct_above60']]
                for col, v in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r, column=col, value=v)
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if col == 2:
                        cell.font = BOLD
                    elif col >= 5:
                        cell.font = BODY_FONT; cell.fill = DATA_FILL
                    else:
                        cell.font = BODY_FONT
                    if band and col in (2, 3):
                        cell.fill = band
                if comp_label == 'Annual':
                    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
                    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=2)
                    ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=3)
                r += 1

    r += 2
    ws.cell(row=r, column=1, value=coordinator_name).font = SIGN_FONT
    ws.cell(row=r, column=1).border = Border(top=Side(style='thin'))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value=principal_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=n)
    r += 1
    ws.cell(row=r, column=1, value=coordinator_role).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value='Principal').font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=n)
    r += 1
    ws.cell(row=r, column=1, value=institute_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=6, value=institute_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=n)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 13
    for col_letter in 'EFGHIJ':
        ws.column_dimensions[col_letter].width = 9
    ws.sheet_view.showGridLines = False
    apply_print_setup(ws, n, last_row_for_header=hr)
    return ws


def build_j12_sheet(wb, sheet_title, sections, institute_name, session_text,
                     academic_year, coordinator_name, coordinator_role, principal_name,
                     code_map=None):
    """J12-style proforma: one row per HEAD (Theory and Practical), TM =
    Internal-component score index, TH = External-component score index,
    Difference = TM - TH, coloured pink when it exceeds 20."""
    code_map = code_map or {}
    form_code = f"{sections[0][1]['scheme']}12"
    ws = wb.create_sheet(f'{sheet_title} {form_code}'[:31])

    headers = ['Sr.\nNo.', 'Course Name', 'TM\n(Sessional Average)\nScore index',
               'TH\n(Annual Exam)\nScore index', 'Difference', 'Remarks\n(if difference > 20%)']
    n = len(headers)

    c = ws.cell(row=1, column=n, value=form_code)
    c.font = Font(name='Calibri', italic=True, bold=True, size=11)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
    c = ws.cell(row=3, column=1, value='Maharashtra State Board of Technical Education, Mumbai')
    c.font = Font(name='Calibri', bold=True, size=16, color='FFFFFF'); c.fill = BANNER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 26

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n)
    c = ws.cell(row=4, column=1, value='SESSIONAL THEORY MARKS (TM) AND ANNUAL THEORY MARKS (TH) ANALYSIS')
    c.font = Font(name='Calibri', bold=True, size=13, color='1F3864'); c.fill = SUBTITLE_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 30

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n)
    c = ws.cell(row=5, column=1, value='Theory Sessional and Theory Annual Examination Result Analysis')
    c.font = Font(name='Calibri', italic=True, size=11); c.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=n)
    c = ws.cell(row=6, column=1, value=institute_name)
    c.font = Font(name='Calibri', bold=True, size=12); c.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    ws.cell(row=7, column=1, value='Academic Year:').font = BOLD
    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=3)
    ws.cell(row=7, column=3, value=academic_year)
    ws.merge_cells(start_row=7, start_column=4, end_row=7, end_column=4)
    ws.cell(row=7, column=4, value='Examination:').font = BOLD
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=n)
    ws.cell(row=7, column=5, value=session_text)
    ws.row_dimensions[7].height = 18
    for col in range(1, n + 1):
        ws.cell(row=7, column=col).border = BORDER
        ws.cell(row=7, column=col).alignment = Alignment(horizontal='left', vertical='center')

    hr = 9
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=i, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[hr].height = 58

    r = hr + 1
    sr = 1
    for section_label, block in sections:
        cols = block['cols']
        students = block['students']
        ext_stats = analytics(students, cols, 'ext')
        inn_stats = analytics(students, cols, 'inn')
        for c in cols:
            tm = inn_stats[c['n']]['score_index']
            th = ext_stats[c['n']]['score_index']
            diff = tm - th
            ws.cell(row=r, column=1, value=sr)
            code = code_map.get(c['subject'], '')
            name_cell = ws.cell(row=r, column=2, value=f"{c['subject']} – {code}")
            name_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=r, column=3, value=tm)
            ws.cell(row=r, column=4, value=th)
            diff_cell = ws.cell(row=r, column=5, value=diff)
            remark_cell = ws.cell(row=r, column=6, value='Difference > 20%' if diff > 20 else '')
            fill = DIFF_PINK if diff > 20 else DIFF_GREEN
            for col in range(1, n + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER
                cell.font = BODY_FONT
                if col != 2:
                    cell.alignment = CENTER
            diff_cell.fill = fill
            remark_cell.fill = fill
            r += 1
            sr += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    note = ws.cell(row=r, column=1,
                    value='Score Index = (Total Marks Secured by All Students / Total of Maximum Marks for the Head) × 100')
    note.font = Font(name='Calibri', italic=True, size=9)
    r += 2

    ws.cell(row=r, column=1, value=coordinator_name).font = SIGN_FONT
    ws.cell(row=r, column=1).border = Border(top=Side(style='thin'))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=4, value=principal_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=n)
    r += 1
    ws.cell(row=r, column=1, value=coordinator_role).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=4, value='Principal').font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=n)
    r += 1
    ws.cell(row=r, column=1, value=institute_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=4, value=institute_name).font = SIGN_FONT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=n)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 20
    ws.sheet_view.showGridLines = False
    apply_print_setup(ws, n, last_row_for_header=hr)
    return ws


def build_notes_sheet(wb, lines: list[str]):
    ws = wb.create_sheet('Notes & Assumptions', 0)
    ws.column_dimensions['A'].width = 100
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name='Calibri', size=11, bold=(i == 1))
    return ws
